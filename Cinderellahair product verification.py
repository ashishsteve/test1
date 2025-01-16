import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl

url = 'http://dev21.gmrwebsites.com/Account/Login'


ID_xpath = '//*[@id="UserName"]'
Pass_xpath = '//*[@id="myInput"]'
Login_xpath = '//*[@id="loginform"]/form/div[5]/input'

chrome_options = Options()
#chrome_options.add_argument("--headless")
driver = webdriver.Chrome(options=chrome_options)
driver.get(url)
wait = WebDriverWait(driver, 10)

ID_element = wait.until(EC.presence_of_element_located((By.XPATH, ID_xpath)))
Pass_element = wait.until(EC.presence_of_element_located((By.XPATH, Pass_xpath)))
Login_element = wait.until(EC.element_to_be_clickable((By.XPATH, Login_xpath)))

ID_element.send_keys("admin@cinderellahair.com")
Pass_element.send_keys("c!inderell@")

Login_element.click()
time.sleep(5)

Product_xpath = '//*[@id="sidebar-menu"]/div/ul/li[4]/a'
Product_xpath1 = wait.until(EC.presence_of_element_located((By.XPATH, Product_xpath)))
Product_xpath1.click()
time.sleep(2)

Product_list = '//*[@id="sidebar-menu"]/div/ul/li[4]/ul/li[1]/a'
Product_list1 = wait.until(EC.presence_of_element_located((By.XPATH, Product_list)))
Product_list1.click()
time.sleep(5)


file_path = "C:\\Users\\ashish sinha\\desktop\\Cinderellahair Products Web January 2025.xlsx"

file = file_path
wb = openpyxl.load_workbook(file)
sheet = wb["Sheet1"]

time.sleep(1)
for i, row in enumerate(sheet.iter_rows(min_row=2, min_col=2, values_only=True), start=2):

    ITEM_NO = row[0]
    print(str(ITEM_NO))

    Search = '//*[@id="Name"]'
    Search1 = wait.until(EC.presence_of_element_located((By.XPATH, Search)))
    Search1.clear()
    Search1.send_keys(ITEM_NO)

    Click = '//*[@id="search"]/div[2]/button'
    Click1 = wait.until(EC.presence_of_element_located((By.XPATH, Click)))
    Click1.click()
    time.sleep(5)

    Price_xpath1 = '//*[@id="paginationdiv"]/div[3]/div/div/div[2]/table/tbody/tr[1]/td[5]'
    Price_Xpath2 = '//*[@id="paginationdiv"]/div[3]/div/div/div[2]/table/tbody/tr/td[5]'

    try:
        Price1 = wait.until(EC.presence_of_element_located((By.XPATH, Price_xpath1)))
        print("Price:- ", Price1.text)
        sheet.cell(row=i, column=7).value = Price1.text
    except :
        try:
            Price2 = wait.until(EC.presence_of_element_located((By.XPATH, Price_Xpath2)))
            print("Price:- ", Price2.text)
            sheet.cell(row=i, column=7).value = Price2.text
        except:
            print("Price not found.")

            print()



        wb.save(file_path)
        driver.quit()


